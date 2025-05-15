import openpyxl

txtfiles = [
    "./english.txt",
    "./schinese.txt",
    "./tchinese.txt",
    "./japanese.txt",
    "./KOERA.txt",
    "./russian.txt",
    "./turkey.txt",
    "./TAI.txt",
    "./czech.txt",
    "./spanish.txt",
    "./UKRAINIAN.txt",
    "./german.txt",
    "./french.txt",
    "./italian.txt",
    "./portuguese.txt",
    "./DUTCH.txt",
    "./hebrew.txt",
    "./polish.txt",
]
file_path = r"./font.xlsx"  # 替换为你的文件路径
workbook = openpyxl.load_workbook(file_path)
# 2. 选择工作表（按名称）
sheet = workbook["Sheet1"]  # 替换为你的工作表名
rowcount = 0
# 遍历第一列（A列）所有单元格
for cell in sheet["A"]:
    if cell.value is not None and str(cell.value).strip() != "":
        rowcount += 1
rowcount += 2
with open("../font.tab", "w", encoding="utf-8") as f:
    for row in range(2, rowcount):
        f.write(sheet.cell(row=row, column=1).value + "\n")
for file in txtfiles:
    with open(file, "w", encoding="utf-8-sig") as f:
        for row in range(2, rowcount):
            if sheet.cell(row=row, column=txtfiles.index(file) + 2).value is None:
                f.write(" ")
            else:
                f.write(
                    sheet.cell(row=row, column=txtfiles.index(file) + 2).value + "\n"
                )
