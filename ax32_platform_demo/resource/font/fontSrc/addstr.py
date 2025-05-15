import requests
import hashlib
import random
import time
import openpyxl
from openpyxl.styles import Font
from translate import Translator

# 1. 读取现有文件
file_path = r"./font.xlsx"  # 替换为你的文件路径


class BaiduTranslator:
    def __init__(self, app_id, secret_key):
        self.app_id = app_id
        self.secret_key = secret_key
        self.base_url = "https://fanyi-api.baidu.com/api/trans/vip/translate"

        # 严格顺序的语言列表（百度官方代码 -> 显示名称）
        self.ordered_langs = [
            ("en", "英语(english)"),
            ("zh", "简体中文(schinese)"),
            ("cht", "繁体中文(tchinese)"),
            ("jp", "日语(japanese)"),
            ("kor", "韩语(koera)"),
            ("ru", "俄语(russian)"),
            ("tr", "土耳其语(turkey)"),
            ("th", "泰语(tai)"),
            ("cs", "捷克语(czech)"),
            ("spa", "西班牙语(spanish)"),
            ("uk", "乌克兰语(ukrainian)"),
            ("de", "德语(german)"),
            ("fra", "法语(french)"),
            ("it", "意大利语(italian)"),
            ("pt", "葡萄牙语(portuguese)"),
            ("nl", "荷兰语(dutch)"),
            ("he", "希伯来语(hebrew)"),
            ("pl", "波兰语(polish)"),
        ]

    def _generate_signature(self, text, salt):
        """生成符合百度要求的MD5签名"""
        sign_str = f"{self.app_id}{text}{salt}{self.secret_key}"
        return hashlib.md5(sign_str.encode()).hexdigest()

    def _split_text(self, text, chunk_size=2000):
        """处理长文本分块（API限制单次请求不超过2000字符）"""
        return [text[i : i + chunk_size] for i in range(0, len(text), chunk_size)]

    def _api_request(self, text, target_lang):
        """执行API请求并处理响应"""
        salt = str(random.randint(10000, 99999))
        params = {
            "q": text,
            "from": "zh",
            "to": target_lang,
            "appid": self.app_id,
            "salt": salt,
            "sign": self._generate_signature(text, salt),
        }

        try:
            response = requests.post(self.base_url, params=params, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"网络请求失败: {str(e)}")
            return None

    def translate(self, text, target_lang, retries=3):
        """带自动重试的翻译核心方法"""
        chunks = self._split_text(text)
        result = []

        for chunk in chunks:
            for attempt in range(retries):
                response = self._api_request(chunk, target_lang)
                if response and "trans_result" in response:
                    result.append(response["trans_result"][0]["dst"])
                    break
            else:
                return None

        return "".join(result)

    def batch_translate(self, text):
        """保持顺序的批量翻译方法"""
        results = []
        for lang_code, lang_name in self.ordered_langs:
            if lang_code == "tr" or lang_code == "uk" or lang_code == "he":
                translator = Translator(to_lang=lang_code, from_lang="zh-CN")
                results.append(translator.translate(text))
            else:
                translated = self.translate(text, lang_code)
                if lang_code == "en":
                    translated = translated.title()
                results.append(translated)
            print(lang_name + " translate success " + results[-1])
        return results


# ==================== 使用示例 ====================
if __name__ == "__main__":
    # 配置信息（从百度控制台获取）
    APP_ID = "20250429002345436"  # 替换为实际值
    SECRET_KEY = "PwvnO9CAMIM6e2tI6pCT"  # 替换为实际值
    custom_font1 = Font(
        name="宋体",  # 字体名称
        size=11,  # 字号
        bold=False,  # 加粗
        italic=False,  # 斜体
        color="000000",  # 颜色（十六进制或colors模块预定义值）
    )
    custom_font = Font(
        name="华文宋体",  # 字体名称
        size=11,  # 字号
        bold=False,  # 加粗
        italic=False,  # 斜体
        color="000000",  # 颜色（十六进制或colors模块预定义值）
    )
    workbook = openpyxl.load_workbook(file_path)
    # 2. 选择工作表（按名称）
    sheet = workbook["Sheet1"]  # 替换为你的工作表名
    rowcount = 0
    # 遍历第一列（A列）所有单元格
    for cell in sheet["A"]:
        if cell.value is not None and str(cell.value).strip() != "":
            rowcount += 1
    rowcount += 2
    user_input = input("\n请输入要翻译的中文文本: ").strip()
    key_name = input("请输入键名（如STR_NEW_KEY）: ").strip()
    sheet.cell(row=rowcount, column=1, value=(key_name + ",")).font = custom_font1
    # 初始化翻译器
    translator = BaiduTranslator(APP_ID, SECRET_KEY)
    # 执行批量翻译
    start_time = time.time()
    translations = translator.batch_translate(user_input)

    # 输出结果
    i = 1
    for trans in translations:
        i += 1
        sheet.cell(row=rowcount, column=i, value=('"' + trans + '"')).font = custom_font
        print(trans)
    print(f"\n总耗时: {time.time()-start_time:.2f}秒")
    # 4. 保存文件
    workbook.save(file_path)  # 覆盖原文件
